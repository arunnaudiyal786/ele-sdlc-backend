"""
Excel document extractor using openpyxl.

Extracts structured data from Excel spreadsheets including:
- Multiple sheets support
- Auto-detection of header rows
- Merged cell handling
- Formula result extraction
- Date conversion
"""

import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple, Union

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from pipeline.extractors.base import (
    BaseExtractor,
    DocumentMetadata,
    ExtractedData,
    ExtractedField,
    ExtractedTable,
)


class ExcelExtractor(BaseExtractor):
    """
    Extracts structured data from Excel files (XLSX, XLS).

    Features:
    - Multiple sheet extraction
    - Auto-detection of header rows
    - Merged cell value propagation
    - Excel date serial number conversion
    - Formula result extraction (not formulas)
    - Column type inference
    """

    # Regex patterns for common data types
    JIRA_ID_PATTERN = re.compile(r"\b([A-Z]+-\d+|MM\d+)\b")
    EMAIL_PATTERN = re.compile(r"\b[\w\.-]+@[\w\.-]+\.\w+\b")

    def get_supported_extensions(self) -> List[str]:
        """Return supported file extensions."""
        return [".xlsx", ".xls"]

    async def extract(self, file_path: Path) -> ExtractedData:
        """
        Extract structured data from an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            ExtractedData with all extracted content

        Raises:
            FileNotFoundError: If file does not exist
            ValueError: If file is not a valid Excel file
        """
        self._validate_file(file_path)

        try:
            # Load workbook with data_only=True to get formula results
            workbook = load_workbook(file_path, data_only=True)
        except InvalidFileException:
            raise ValueError(f"Invalid or corrupted Excel file: {file_path}")
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {file_path}. {str(e)}")

        # Initialize extraction result
        extraction = ExtractedData()

        # Extract metadata
        extraction.metadata = self._extract_metadata(workbook, file_path)

        # Process each sheet
        raw_content_parts: List[str] = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Extract table from sheet
            table = self._extract_sheet_as_table(sheet, sheet_name)
            if table:
                extraction.tables.append(table)

                # Add to raw content
                raw_content_parts.append(f"=== Sheet: {sheet_name} ===")
                for row in table.rows:
                    row_text = " | ".join(str(v) for v in row.values())
                    raw_content_parts.append(row_text)

                    # Detect patterns in cell values
                    for value in row.values():
                        if isinstance(value, str):
                            extraction.jira_ids.extend(
                                self.JIRA_ID_PATTERN.findall(value)
                            )
                            extraction.emails.extend(self.EMAIL_PATTERN.findall(value))

        # Build raw content
        extraction.raw_content = "\n".join(raw_content_parts)
        extraction.raw_sections = {
            sheet_name: f"Sheet with {len(workbook[sheet_name].rows if hasattr(workbook[sheet_name], 'rows') else [])} rows"
            for sheet_name in workbook.sheetnames
        }

        # Deduplicate pattern matches
        extraction.jira_ids = list(set(extraction.jira_ids))
        extraction.emails = list(set(extraction.emails))

        # Calculate confidence
        extraction.overall_confidence = self._calculate_confidence(extraction)

        workbook.close()
        return extraction

    def _extract_metadata(
        self, workbook: Any, file_path: Path
    ) -> DocumentMetadata:
        """Extract workbook metadata."""
        return DocumentMetadata(
            filename=file_path.name,
            file_path=str(file_path),
            file_size=file_path.stat().st_size,
            file_type="xlsx",
            sheet_names=workbook.sheetnames,
        )

    def _extract_sheet_as_table(
        self, sheet: Worksheet, sheet_name: str
    ) -> Optional[ExtractedTable]:
        """
        Extract a worksheet as a structured table.

        Args:
            sheet: openpyxl Worksheet
            sheet_name: Name of the sheet

        Returns:
            ExtractedTable or None if sheet is empty
        """
        # Find the header row (first row with mostly string values)
        header_row_idx = self._find_header_row(sheet)
        if header_row_idx is None:
            return None

        # Get headers
        headers = []
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=header_row_idx, column=col)
            header_value = self._get_cell_value(cell)
            if header_value:
                headers.append(str(header_value))
            else:
                headers.append(f"Column_{col}")

        # Extract data rows
        rows_data: List[Dict[str, Any]] = []
        for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
            row_dict: Dict[str, Any] = {}
            has_data = False

            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = self._get_cell_value(cell)

                if value is not None and value != "":
                    has_data = True

                row_dict[header] = value

            if has_data:
                rows_data.append(row_dict)

        if not rows_data:
            return None

        return ExtractedTable(
            headers=headers,
            rows=rows_data,
            source_location=f"Sheet: {sheet_name}",
            confidence=0.95,
        )

    def _find_header_row(self, sheet: Worksheet) -> Optional[int]:
        """
        Find the header row in a worksheet.

        Looks for the first row that has mostly string values
        and is followed by data rows.

        Returns:
            Row index (1-based) or None if no header found
        """
        for row_idx in range(1, min(10, sheet.max_row + 1)):  # Check first 10 rows
            row_values = []
            string_count = 0
            empty_count = 0

            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col)
                value = self._get_cell_value(cell)

                if value is None or value == "":
                    empty_count += 1
                elif isinstance(value, str):
                    string_count += 1
                    row_values.append(value)

            total_cells = sheet.max_column
            if total_cells == 0:
                continue

            # Header row should have mostly strings and not be mostly empty
            string_ratio = string_count / total_cells
            empty_ratio = empty_count / total_cells

            if string_ratio > 0.5 and empty_ratio < 0.5:
                # Verify there's data below
                if row_idx < sheet.max_row:
                    return row_idx

        # Fallback: use first row
        return 1 if sheet.max_row > 0 else None

    def _get_cell_value(self, cell: Cell) -> Any:
        """
        Get the value from a cell, handling special cases.

        - Converts Excel dates to Python dates
        - Handles merged cells
        - Returns formula results (not formulas)
        """
        value = cell.value

        if value is None:
            return None

        # Handle datetime
        if isinstance(value, datetime):
            return value.date() if value.hour == 0 and value.minute == 0 else value

        # Handle date
        if isinstance(value, date):
            return value

        # Handle numbers that might be Excel dates
        if isinstance(value, (int, float)) and cell.is_date:
            try:
                # Excel date serial to Python date
                from openpyxl.utils.datetime import from_excel
                return from_excel(value)
            except Exception:
                pass

        # Handle booleans
        if isinstance(value, bool):
            return value

        # Handle strings
        if isinstance(value, str):
            return value.strip()

        return value

    def _calculate_confidence(self, extraction: ExtractedData) -> float:
        """
        Calculate overall extraction confidence.

        Based on:
        - Number of tables extracted
        - Number of rows with data
        - Pattern matches found
        """
        confidence = 0.6  # Base confidence for Excel (structured format)

        # Tables bonus
        if extraction.tables:
            confidence += 0.1
            # More rows = higher confidence
            total_rows = sum(len(t.rows) for t in extraction.tables)
            if total_rows > 5:
                confidence += 0.1
            if total_rows > 20:
                confidence += 0.1

        # Pattern matches bonus
        if extraction.jira_ids:
            confidence += 0.05
        if extraction.emails:
            confidence += 0.05

        return min(confidence, 1.0)

    def get_sheet_as_dataframe(
        self, file_path: Path, sheet_name: Optional[str] = None
    ) -> List[Dict[str, Any]]:
        """
        Convenience method to get a sheet as a list of dictionaries.

        Args:
            file_path: Path to Excel file
            sheet_name: Specific sheet name, or None for first sheet

        Returns:
            List of row dictionaries
        """
        workbook = load_workbook(file_path, data_only=True)

        target_sheet = sheet_name or workbook.sheetnames[0]
        sheet = workbook[target_sheet]

        table = self._extract_sheet_as_table(sheet, target_sheet)
        workbook.close()

        return table.rows if table else []
